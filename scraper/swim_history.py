"""
MAAC Swim Club - Full Swim History Scraper
Pulls every recorded swim per event per swimmer, including split times.

Pipeline position:
    maac_scraper.py → [swim_history.py] → build_tables.py

Input:
    ../data/maac_best_times.csv   (from maac_scraper.py)

Output:
    ../data/maac_swim_history.xlsx
    ../data/maac_swim_history.csv

Requirements:
    pip install requests pandas openpyxl python-dotenv

Setup:
    Create scraper/.env with your SwimCloud session cookie:
        SWIMCLOUD_SESSION=your_session_cookie_here

Usage:
    cd scraper
    python swim_history.py

Note:
    This script makes ~1 API call per unique (distance, course) per swimmer.
    With 132 swimmers averaging ~8 distance/course combos each, expect 15-30 minutes.
    Do not run this unnecessarily — only needed when refreshing full history.

API note:
    SwimCloud's times_by_event endpoint ignores the stroke parameter and returns
    ALL swims at a given distance/course for a swimmer. Stroke is assigned by
    matching each returned swim time to the swimmer's closest best time per stroke.
"""

import sys
import os
import time
import pandas as pd
from dotenv import load_dotenv
sys.stdout.reconfigure(encoding="utf-8")
from curl_cffi import requests

# ── Paths ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR   = os.path.join(SCRIPT_DIR, "..", "data")
os.makedirs(DATA_DIR, exist_ok=True)

# ── Config ────────────────────────────────────────────────────────────────────

load_dotenv()
SESSION_COOKIE = os.getenv("SWIMCLOUD_SESSION")

if not SESSION_COOKIE:
    raise SystemExit(
        "\n❌ SWIMCLOUD_SESSION not found in .env file.\n"
        "   Create scraper/.env and add:\n"
        "   SWIMCLOUD_SESSION=your_session_cookie_here\n"
        "   (Log into swimcloud.com → F12 → Application → Cookies → copy sessionid)"
    )

BASE_URL = "https://www.swimcloud.com"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer":          "https://www.swimcloud.com/",
    "Accept":           "application/json, text/plain, */*",
    "Accept-Language":  "en-US,en;q=0.9",
    "X-Requested-With": "XMLHttpRequest",
}

STROKE_TO_CODE = {"Free": "1", "Back": "2", "Breast": "3", "Fly": "4", "IM": "5"}
COURSE_TO_CODE = {"SCY": "Y", "LCM": "L", "SCM": "S"}

# Reverse map: API single-letter code → our course name (or None to discard)
API_COURSE_TO_NAME  = {"Y": "SCY", "L": "LCM", "S": None, "M": None}
TRACKED_COURSES     = {"SCY", "LCM"}

# Reverse map: API eventstroke code → our stroke name
API_STROKE_TO_NAME  = {"1": "Free", "2": "Back", "3": "Breast", "4": "Fly", "5": "IM"}

# Approximate world records in seconds, keyed by (stroke, distance, course).
# Used as a sanity floor — no college swimmer should be faster than these.
# Times set slightly below actual WRs to avoid false positives on data entry quirks.
_WR = {
    # SCY men / women (same table — men's WRs used as absolute floor)
    ("Free",   25,  "SCY"): 9.0,
    ("Free",   50,  "SCY"): 17.0,
    ("Free",  100,  "SCY"): 39.0,
    ("Free",  200,  "SCY"): 86.0,
    ("Free",  400,  "SCY"): 188.0,
    ("Free",  500,  "SCY"): 235.0,
    ("Free", 1000,  "SCY"): 480.0,
    ("Free", 1650,  "SCY"): 800.0,
    ("Back",   50,  "SCY"): 20.0,
    ("Back",  100,  "SCY"): 43.0,
    ("Back",  200,  "SCY"): 95.0,
    ("Breast", 50,  "SCY"): 22.0,
    ("Breast",100,  "SCY"): 49.0,
    ("Breast",200,  "SCY"): 110.0,
    ("Fly",    50,  "SCY"): 19.0,
    ("Fly",   100,  "SCY"): 42.0,
    ("Fly",   200,  "SCY"): 95.0,
    ("IM",    100,  "SCY"): 46.0,
    ("IM",    200,  "SCY"): 95.0,
    ("IM",    400,  "SCY"): 200.0,
    # LCM
    ("Free",   50,  "LCM"): 20.0,
    ("Free",  100,  "LCM"): 46.0,
    ("Free",  200,  "LCM"): 102.0,
    ("Free",  400,  "LCM"): 220.0,
    ("Free",  800,  "LCM"): 455.0,
    ("Free", 1500,  "LCM"): 860.0,
    ("Back",   50,  "LCM"): 23.0,
    ("Back",  100,  "LCM"): 51.0,
    ("Back",  200,  "LCM"): 113.0,
    ("Breast", 50,  "LCM"): 25.0,
    ("Breast",100,  "LCM"): 56.0,
    ("Breast",200,  "LCM"): 125.0,
    ("Fly",    50,  "LCM"): 22.0,
    ("Fly",   100,  "LCM"): 49.0,
    ("Fly",   200,  "LCM"): 111.0,
    ("IM",    200,  "LCM"): 113.0,
    ("IM",    400,  "LCM"): 240.0,
}
# Ceiling multiplier: times slower than WR × this value are flagged as garbled
WR_CEILING_MULTIPLIER = 4.0


def validate_time(swim_secs, stroke, distance, course):
    """
    Check a swim time against world record bounds.
    Returns (valid: bool, reason: str | None).
    """
    wr = _WR.get((stroke, distance, course))
    if wr is None:
        return True, None  # no reference for this event — pass through
    if swim_secs < wr:
        return False, f"faster than WR floor ({wr}s)"
    if swim_secs > wr * WR_CEILING_MULTIPLIER:
        return False, f"slower than {WR_CEILING_MULTIPLIER}× WR ({wr * WR_CEILING_MULTIPLIER:.1f}s)"
    return True, None

# Stroke assignment confidence thresholds.
# Margin is measured as the difference in *percentage slower than best* between
# the closest and second-closest stroke. Using percentage rather than raw seconds
# accounts for swimmers improving over time — an old swim may be several seconds
# from the current best, so absolute margins are unreliable.
#
# Example: if a swim is 8% slower than the Free best but 15% slower than the Back
# best, the margin is 7pp. That's a confident Free assignment regardless of distance.
#
# Below SKIP_THRESHOLD   → too ambiguous to trust; record is dropped
# Below FLAG_THRESHOLD   → kept but marked confidence="low" for review
STROKE_SKIP_THRESHOLD = 3.0   # percentage points
STROKE_FLAG_THRESHOLD = 10.0  # percentage points


# ── Session ───────────────────────────────────────────────────────────────────

def build_session():
    s = requests.Session(impersonate="chrome124")
    s.headers.update(HEADERS)
    s.cookies.set("sessionid", SESSION_COOKIE, domain="swimcloud.com")
    s.cookies.set("sessionid", SESSION_COOKIE, domain="www.swimcloud.com")
    cf = os.getenv("SWIMCLOUD_CF_CLEARANCE")
    if cf:
        s.cookies.set("cf_clearance", cf, domain="swimcloud.com")
        s.cookies.set("cf_clearance", cf, domain="www.swimcloud.com")
    return s

SESSION = build_session()


def get(url, retries=3):
    """GET with retry. Exits immediately on 403."""
    for attempt in range(retries):
        try:
            resp = SESSION.get(url, timeout=15)

            if resp.status_code == 200:
                return resp

            elif resp.status_code == 403:
                raise SystemExit(
                    "\n❌ 403 Forbidden — session cookie has expired.\n"
                    "   Get a fresh one:\n"
                    "     1. Log into swimcloud.com in Chrome\n"
                    "     2. F12 → Application → Cookies → www.swimcloud.com\n"
                    "     3. Copy 'sessionid' value into scraper/.env"
                )

            elif resp.status_code == 429:
                wait = 20 * (attempt + 1)
                print(f"\n  ⚠ Rate limited — waiting {wait}s...", end="", flush=True)
                time.sleep(wait)

            else:
                wait = 3 * (attempt + 1)
                print(f"\n  ⚠ HTTP {resp.status_code} — retrying in {wait}s...", end="", flush=True)
                time.sleep(wait)

        except SystemExit:
            raise
        except Exception as e:
            wait = 3 * (attempt + 1)
            print(f"\n  ⚠ Error: {e} — retrying in {wait}s...", end="", flush=True)
            time.sleep(wait)

    return None


# ── Time helpers ──────────────────────────────────────────────────────────────

def time_to_seconds(t):
    """Convert 'M:SS.ss' or 'SS.ss' string to float seconds. Returns None on failure."""
    try:
        s = str(t).strip()
        if ":" in s:
            m, sec = s.split(":", 1)
            return float(m) * 60 + float(sec)
        return float(s)
    except Exception:
        return None


def assign_stroke(swim_time_str, stroke_bests):
    """
    Assign a stroke to a swim by finding which stroke's best time is closest.

    stroke_bests: dict of {stroke: best_time_str}
    Returns (stroke, margin) where margin is the gap in seconds between the
    closest and second-closest stroke best time. margin=inf means only one
    stroke candidate existed (unambiguous). Returns (None, None) on failure.

    Because the SwimCloud times_by_event API ignores the stroke parameter and
    returns all swims at a given distance/course, we use this heuristic to
    recover the correct stroke label.
    """
    swim_secs = time_to_seconds(swim_time_str)
    if swim_secs is None:
        return None, None

    # Score each stroke as % slower than its best time.
    # A swim faster than the best is physically impossible for that stroke — exclude it.
    diffs = []
    for stroke, best_t in stroke_bests.items():
        best_secs = time_to_seconds(best_t)
        if best_secs is None or best_secs <= 0:
            continue
        pct_slower = (swim_secs - best_secs) / best_secs * 100
        if pct_slower < 0:
            continue  # faster than current best — can't be this stroke
        diffs.append((pct_slower, stroke))

    if not diffs:
        return None, None

    diffs.sort()
    best_stroke = diffs[0][1]
    margin = (diffs[1][0] - diffs[0][0]) if len(diffs) > 1 else float("inf")
    return best_stroke, margin


# ── Event history ─────────────────────────────────────────────────────────────

def get_swims_for_event(swimmer_id, stroke, distance, course):
    """
    Fetch all recorded swims for a swimmer at a specific stroke/distance/course.
    The API filters by stroke code, so each stroke requires its own call.
    """
    course_code  = COURSE_TO_CODE.get(course)
    stroke_code  = STROKE_TO_CODE.get(stroke)
    if not course_code or not stroke_code:
        return []

    event_param = f"{stroke_code}|{distance}|{course_code}|1"
    url = (
        f"{BASE_URL}/api/swimmers/{swimmer_id}/times_by_event/"
        f"?event={requests.utils.quote(event_param)}"
    )

    resp = get(url)
    if resp is None:
        return []

    try:
        data = resp.json()
    except Exception:
        return []

    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ["times", "results", "data", "swims"]:
            if key in data and isinstance(data[key], list):
                return data[key]
    return []


# ── Export ────────────────────────────────────────────────────────────────────

def export(all_records):
    df = pd.DataFrame(all_records)
    df = df[[
        "name", "gender", "swimmer_id", "distance", "stroke", "course",
        "time", "meet", "date", "place", "heat", "lane", "splits", "confidence"
    ]]
    df = df.sort_values(
        ["gender", "name", "course", "distance", "stroke", "date"]
    ).reset_index(drop=True)

    # CSV
    csv_path = os.path.join(DATA_DIR, "maac_swim_history.csv")
    df.to_csv(csv_path, index=False)
    n_low  = (df["confidence"] == "low").sum()
    n_high = (df["confidence"] == "high").sum()
    print(f"   ✅ data/maac_swim_history.csv — {len(df):,} records "
          f"({n_high:,} high-confidence, {n_low:,} low-confidence)")

    # Flagged assignments report
    flagged = df[df["confidence"] == "low"]
    if not flagged.empty:
        flag_path = os.path.join(DATA_DIR, "flagged_assignments.csv")
        flagged.to_csv(flag_path, index=False)
        print(f"   ⚠  data/flagged_assignments.csv — {len(flagged):,} low-confidence rows for review")

    # Excel — times forced as text
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Swim History"
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append(list(row))

        time_col = list(df.columns).index("time") + 1
        for row in ws.iter_rows(min_row=2, min_col=time_col, max_col=time_col):
            for cell in row:
                cell.value         = str(cell.value)
                cell.number_format = "@"

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 2, 40
            )

        # Highlight low-confidence rows in yellow
        from openpyxl.styles import PatternFill
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        conf_col = list(df.columns).index("confidence") + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[conf_col - 1].value == "low":
                for cell in row:
                    cell.fill = yellow

        xlsx_path = os.path.join(DATA_DIR, "maac_swim_history.xlsx")
        wb.save(xlsx_path)
        print(f"   ✅ data/maac_swim_history.xlsx — low-confidence rows highlighted yellow")

    except ImportError:
        print("   ⚠ openpyxl not installed — run: pip install openpyxl")

    return df


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  MAAC SwimCloud Full Swim History Scraper")
    print("=" * 55)
    print("  ⚠  This script takes 15-30 minutes to complete.")
    print("  ⚠  Only run when a full history refresh is needed.")
    print("=" * 55)

    # Load best times to know which events each swimmer has
    best_times_path = os.path.join(DATA_DIR, "maac_best_times.csv")
    if not os.path.exists(best_times_path):
        raise SystemExit(
            f"❌ data/maac_best_times.csv not found.\n"
            f"   Run maac_scraper.py first."
        )

    best_times = pd.read_csv(best_times_path)
    swimmers   = (
        best_times[["name", "swimmer_id", "gender"]]
        .drop_duplicates()
        .to_dict("records")
    )

    # Build best-time lookup: {swimmer_id: {(distance, course): {stroke: best_time}}}
    best_lookup = {}
    for _, row in best_times.iterrows():
        sid  = int(row["swimmer_id"])
        key  = (row["distance"], row["course"])
        best_lookup.setdefault(sid, {}).setdefault(key, {})[row["stroke"]] = row["best_time"]

    print(f"\n📋 {len(swimmers)} swimmers | "
          f"{len(best_times)} events total\n")

    all_records     = []
    total_api_calls = 0
    total_skipped   = 0

    for i, swimmer in enumerate(swimmers, 1):
        name   = swimmer["name"]
        sid    = int(swimmer["swimmer_id"])
        gender = swimmer["gender"]

        stroke_bests_map = best_lookup.get(sid, {})

        # One API call per (stroke, distance, course) event
        event_combos = (
            best_times[best_times["swimmer_id"] == sid][["stroke", "distance", "course"]]
            .drop_duplicates()
            .to_dict("records")
        )

        print(f"\n  [{i:>3}/{len(swimmers)}] {name} ({gender}) "
              f"— {len(event_combos)} events")

        for combo in event_combos:
            stroke   = combo["stroke"]
            distance = combo["distance"]
            course   = combo["course"]
            key      = (distance, course)

            stroke_bests = stroke_bests_map.get(key, {})
            stroke_list  = list(stroke_bests.keys())

            print(f"    {distance} {stroke} {course}...", end="", flush=True)

            entries = get_swims_for_event(str(sid), stroke, distance, course)
            total_api_calls += 1

            # Deduplicate entries by swim ID
            seen_ids = set()
            unique_entries = []
            for e in entries:
                if not isinstance(e, dict):
                    continue
                eid = e.get("id")
                if eid and eid in seen_ids:
                    continue
                if eid:
                    seen_ids.add(eid)
                unique_entries.append(e)

            count   = 0
            skipped = 0
            for entry in unique_entries:
                swim_time = str(entry.get("eventtime", "")).strip()
                if not swim_time or swim_time == "None":
                    continue

                # Drop DQ / NS / scratch — legal=False or timecode set
                if entry.get("legal") is False:
                    continue
                if entry.get("timecode"):
                    continue

                # Filter by course using the eventcourse field — drop SCM and unknowns.
                raw_course_code = entry.get("eventcourse")
                if raw_course_code is not None:
                    entry_course = API_COURSE_TO_NAME.get(str(raw_course_code).upper())
                    if entry_course not in TRACKED_COURSES:
                        continue  # SCM or unknown — discard
                else:
                    entry_course = course  # field absent; trust the query

                # Assign stroke directly from eventstroke field.
                # No heuristic needed — the API tags each entry with its actual stroke.
                raw_stroke_code = entry.get("eventstroke")
                if raw_stroke_code is not None:
                    entry_stroke = API_STROKE_TO_NAME.get(str(raw_stroke_code))
                    if entry_stroke is None:
                        continue  # unrecognised stroke code — skip
                    confidence = "high"
                else:
                    # Fallback: heuristic if field absent
                    entry_stroke, margin = assign_stroke(swim_time, {stroke: stroke_bests[stroke]} if stroke in stroke_bests else stroke_bests)
                    if entry_stroke is None:
                        continue
                    if len(stroke_bests) > 1 and margin < STROKE_SKIP_THRESHOLD:
                        skipped += 1
                        continue
                    confidence = "high" if (len(stroke_bests) == 1 or margin >= STROKE_FLAG_THRESHOLD) else "low"

                # Validate time against world record bounds
                swim_secs = time_to_seconds(swim_time)
                if swim_secs is None or swim_secs <= 0:
                    continue
                valid, reason = validate_time(swim_secs, stroke, distance, entry_course)
                if not valid:
                    print(f"\n  [SKIP] {name} {distance} {entry_stroke} {entry_course} "
                          f"{swim_time} on {entry.get('dateofswim','?')} — {reason}")
                    continue

                split_data  = entry.get("split", {})
                splits_list = split_data.get("normalized_splittimes", []) if split_data else []
                splits      = ", ".join(str(s) for s in splits_list) if splits_list else ""

                all_records.append({
                    "name":       name,
                    "gender":     gender,
                    "swimmer_id": sid,
                    "distance":   distance,
                    "stroke":     entry_stroke,
                    "course":     entry_course,
                    "time":       swim_time,
                    "meet":       str(entry.get("name", "") or entry.get("meet_name", "")).strip(),
                    "date":       str(entry.get("dateofswim", "")).strip(),
                    "place":      str(entry.get("place", "")).strip(),
                    "heat":       entry.get("heat", ""),
                    "lane":       entry.get("lane", ""),
                    "splits":     splits,
                    "confidence": confidence,
                })
                count += 1

            total_skipped += skipped
            suffix = f" {count} swims"
            if skipped:
                suffix += f" ({skipped} skipped — ambiguous)"
            print(suffix if count else " no data")

            time.sleep(1.2)

    print(f"\n{'─' * 55}")
    print(f"  API calls made:    {total_api_calls}")
    print(f"  Records skipped:   {total_skipped} (stroke margin < {STROKE_SKIP_THRESHOLD}s)")
    print(f"  Confidence thresholds: skip <{STROKE_SKIP_THRESHOLD}s | flag <{STROKE_FLAG_THRESHOLD}s")

    if all_records:
        print(f"\n💾 Saving output...")
        df = export(all_records)
        print(f"\n✅ Done — {len(df):,} swim records across {df['name'].nunique()} swimmers")

        # Stroke distribution summary
        print(f"\nStroke distribution:")
        for stroke, cnt in df["stroke"].value_counts().items():
            print(f"  {stroke}: {cnt:,}")
    else:
        print("⚠ No records collected — check your session cookie.")


if __name__ == "__main__":
    main()
